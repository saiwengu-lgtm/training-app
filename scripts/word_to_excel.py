#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""将 Word 试题文档转为 Excel，区分单选和多选题"""
import sys, re
sys.stdout.reconfigure(encoding="utf-8")

from docx import Document
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

INPUT = r"C:\Users\13975\Desktop\保密试题(1).docx"
OUTPUT = r"C:\Users\13975\Desktop\保密试题库.xlsx"

doc = Document(INPUT)
lines = [p.text.strip() for p in doc.paragraphs if p.text.strip()]

# 识别题型分区
single_start = None
multi_start = None
for i, l in enumerate(lines):
    if "单项选择题" in l or "单选" in l:
        single_start = i
    elif "多项选择题" in l or "多选" in l:
        multi_start = i

questions = []
current = None
current_section = "未知"

for line in lines:
    if line.startswith("一、") or line.startswith("二、"):
        if "单项选择" in line or "单选" in line:
            current_section = "单选"
        elif "多项选择" in line or "多选" in line:
            current_section = "多选"
        continue

    # 匹配题号
    m = re.match(r'^(\d+)[、.．]\s*(.*)$', line)
    if m:
        if current:
            questions.append(current)
        current = {
            "num": int(m.group(1)),
            "text": m.group(2) or "",
            "options": [],
            "answer": "",
            "type": current_section
        }
        continue

    # 匹配选项
    m = re.match(r'^([A-Da-d])[.．、]\s*(.*)$', line)
    if m and current:
        current["options"].append(f"{m.group(1).upper()}. {m.group(2)}")
        continue

    # 匹配答案（单选或多选）
    m = re.match(r'^答案[：:]\s*([A-Da-d]+)$', line)
    if m and current:
        current["answer"] = m.group(1).upper()
        continue

# 最后一个
if current:
    questions.append(current)

# 统计
single = [q for q in questions if q["type"] == "单选"]
multi = [q for q in questions if q["type"] == "多选"]
print(f"提取完成：单选题 {len(single)} 题，多选题 {len(multi)} 题，共 {len(questions)} 题")

# 写入 Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "题库"

headers = ["题型", "题目", "选项A", "选项B", "选项C", "选项D", "答案"]
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

body_font = Font(size=11)
wrap_align = Alignment(vertical="center", wrap_text=True)
multi_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

for i, q in enumerate(questions, 2):
    opts = q["options"]
    while len(opts) < 4:
        opts.append("")

    # 题型
    type_cell = ws.cell(row=i, column=1, value=q["type"])
    type_cell.font = Font(size=11, bold=True, color="0066CC" if q["type"] == "单选" else "CC6600")
    type_cell.alignment = Alignment(horizontal="center", vertical="center")
    type_cell.border = thin_border
    if q["type"] == "多选":
        type_cell.fill = multi_fill

    # 题目
    ws.cell(row=i, column=2, value=q["text"]).font = body_font
    ws.cell(row=i, column=2).alignment = wrap_align
    ws.cell(row=i, column=2).border = thin_border

    # 选项
    for j, opt in enumerate(opts[:4], 3):
        cell = ws.cell(row=i, column=j, value=opt)
        cell.font = body_font
        cell.alignment = wrap_align
        cell.border = thin_border

    # 答案
    ans = ws.cell(row=i, column=7, value=q["answer"])
    ans.font = Font(size=11, bold=True, color="C00000")
    ans.alignment = Alignment(horizontal="center", vertical="center")
    ans.border = thin_border

# 列宽
wsc = {1: 8, 2: 50, 3: 30, 4: 30, 5: 30, 6: 30, 7: 10}
for c, w in wsc.items():
    ws.column_dimensions[chr(64 + c)].width = w

wb.save(OUTPUT)
print(f"已保存: {OUTPUT}")
